from typing import List, Optional, Union
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from pydantic import BaseModel, field_serializer
from datetime import datetime
import random
import string
from ..core.database import get_db
from ..models.user import User, UserRole
from ..utils.dependencies import (
    require_system_admin,
    require_school_admin_or_above,
    get_current_user,
    _get_current_user,
    CurrentUserResponse
)

router = APIRouter(prefix="/users", tags=["用户管理"])


class UserCreate(BaseModel):
    username: str
    password: str
    nickname: Optional[str] = None  # 昵称/学生姓名
    role: UserRole
    school_id: Optional[int] = None


class UserUpdate(BaseModel):
    password: Optional[str] = None
    nickname: Optional[str] = None  # 昵称/学生姓名
    role: Optional[UserRole] = None
    school_id: Optional[int] = None


class UserResponse(BaseModel):
    id: int
    username: str
    nickname: Optional[str] = None  # 昵称/学生姓名
    role: str
    school_id: Optional[int] = None
    created_at: Optional[datetime] = None
    is_switched: bool = False  # 是否为切换后的用户
    original_user_id: Optional[int] = None  # 原始管理员用户ID

    @field_serializer('created_at', mode='wrap')
    def serialize_datetime(self, value: Optional[datetime], _info):
        """序列化datetime字段为ISO格式字符串"""
        return value.isoformat() if value else None

    class Config:
        from_attributes = True


@router.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def create_user(
    user_data: UserCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_school_admin_or_above)
):
    """创建用户"""
    # 检查权限：学校管理员只能创建本校用户
    if current_user.role == UserRole.SCHOOL_ADMIN:
        if user_data.school_id != current_user.school_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="只能创建本校用户"
            )
        # 学校管理员不能创建系统管理员
        if user_data.role == UserRole.SYSTEM_ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权创建系统管理员"
            )
    
    # 检查用户名是否已存在
    existing_user = db.query(User).filter(User.username == user_data.username).first()
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="用户名已存在"
        )
    
    from ..utils.security import get_password_hash
    new_user = User(
        username=user_data.username,
        password_hash=get_password_hash(user_data.password),
        nickname=user_data.nickname,  # 添加昵称
        role=user_data.role,
        school_id=user_data.school_id or current_user.school_id
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user


@router.get("", response_model=List[UserResponse])
async def list_users(
    school_id: Optional[int] = None,
    role: Optional[UserRole] = None,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_school_admin_or_above)
):
    """列出用户"""
    query = db.query(User)

    # 学校管理员只能查看本校用户
    if current_user.role == UserRole.SCHOOL_ADMIN:
        query = query.filter(User.school_id == current_user.school_id)
    elif school_id:
        query = query.filter(User.school_id == school_id)

    if role:
        query = query.filter(User.role == role)

    users = query.all()
    return users


@router.get("/template", status_code=status.HTTP_200_OK)
async def download_student_template():
    """下载学生名单模板（Excel格式）"""
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生名单模板"

    # 添加表头
    headers = ["学号", "姓名(昵称)", "密码"]
    ws.append(headers)

    # 添加示例数据
    example_data = [
        ["2024001", "张三", "123456"],
        ["2024002", "李四", "123456"],
        ["2024003", "王五", "123456"],
    ]
    for row in example_data:
        ws.append(row)

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_template.xlsx"}
    )


@router.get("/export", status_code=status.HTTP_200_OK)
async def export_students(
    school_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_school_admin_or_above)
):
    """导出学生名单（Excel格式）"""
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    # 构建查询
    query = db.query(User).filter(User.role == UserRole.STUDENT)

    # 权限控制
    if current_user.role == UserRole.SCHOOL_ADMIN:
        if school_id:
            query = query.filter(User.school_id == school_id)
        else:
            query = query.filter(User.school_id == current_user.school_id)
    elif school_id:
        query = query.filter(User.school_id == school_id)

    students = query.all()

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生名单"

    # 添加表头
    headers = ["学号", "姓名(昵称)", "角色", "学校ID", "创建时间"]
    ws.append(headers)

    # 添加数据
    for student in students:
        row = [
            student.username,
            student.nickname or '-',
            student.role.value,
            student.school_id,
            student.created_at.strftime("%Y-%m-%d %H:%M")
        ]
        ws.append(row)

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"students_{current_user.school_id if current_user.school_id else 'all'}_{current_user.id}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/switch_user", response_model=UserResponse, status_code=status.HTTP_200_OK)
async def switch_user(
    target_user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(_get_current_user)
):
    """系统管理员切换到指定用户"""
    # 检查当前用户是否为系统管理员
    if current_user.role != UserRole.SYSTEM_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有系统管理员可以切换用户"
        )
    # 获取目标用户
    target_user = db.query(User).filter(User.id == target_user_id).first()
    if not target_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="目标用户不存在"
        )

    # 检查目标用户是否为系统管理员
    if target_user.role == UserRole.SYSTEM_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="不能切换到系统管理员"
        )

    # 更新当前用户的切换状态
    current_user.switched_user_id = target_user_id
    current_user.switched_to_username = target_user.username
    current_user.switched_at = datetime.utcnow()
    db.commit()
    db.refresh(current_user)
    db.refresh(target_user)

    # 返回目标用户信息
    # 注意：target_user.created_at 是 datetime 对象，UserResponse 期望 datetime
    return UserResponse(
        id=target_user.id,
        username=target_user.username,
        nickname=target_user.nickname,
        role=target_user.role.value,
        school_id=target_user.school_id,
        created_at=target_user.created_at,
        is_switched=True,
        original_user_id=current_user.id
    )


@router.get("/cancel_switch", response_model=UserResponse, status_code=status.HTTP_200_OK)
async def cancel_switch(
    db: Session = Depends(get_db),
    current_user: User = Depends(_get_current_user)
):
    """取消切换用户，恢复系统管理员身份"""
    # 检查当前用户是否为被切换的用户
    # 如果是，则需要查找原始管理员并取消切换
    admin_user = db.query(User).filter(User.switched_user_id == current_user.id).first()

    if not admin_user:
        # 如果没有找到切换的admin，说明当前就是admin用户
        if current_user.role != UserRole.SYSTEM_ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="只有系统管理员可以取消切换"
            )
        admin_user = current_user
    else:
        # 确认找到的确实是系统管理员
        if admin_user.role != UserRole.SYSTEM_ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="只有系统管理员可以取消切换"
            )

    # 重置切换状态
    admin_user.switched_user_id = None
    admin_user.switched_to_username = None
    admin_user.switched_at = None
    db.commit()
    db.refresh(admin_user)

    # 返回管理员用户信息
    return UserResponse(
        id=admin_user.id,
        username=admin_user.username,
        nickname=admin_user.nickname,
        role=admin_user.role.value,
        school_id=admin_user.school_id,
        created_at=admin_user.created_at,
        is_switched=False,
        original_user_id=None
    )


@router.get("/{user_id}", response_model=UserResponse)
async def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(get_current_user)
):
    """获取用户信息"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )
    
    # 权限检查
    if current_user.role == UserRole.SCHOOL_ADMIN:
        if user.school_id != current_user.school_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权查看其他学校用户"
            )
    elif current_user.role == UserRole.STUDENT:
        if user.id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="只能查看自己的信息"
            )
    
    return user


@router.put("/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user_data: UserUpdate,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_school_admin_or_above)
):
    """更新用户信息"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    # 权限检查
    if current_user.role == UserRole.SCHOOL_ADMIN:
        if user.school_id != current_user.school_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权修改其他学校用户"
            )
        # 学校管理员不能修改角色为系统管理员
        if user_data.role == UserRole.SYSTEM_ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权设置系统管理员角色"
            )

    # 不能修改系统管理员的密码
    if user_data.password and user.role == UserRole.SYSTEM_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="不能修改系统管理员的密码"
        )

    if user_data.password:
        from ..utils.security import get_password_hash
        user.password_hash = get_password_hash(user_data.password)
    if user_data.nickname is not None:
        user.nickname = user_data.nickname
    if user_data.role:
        user.role = user_data.role
    if user_data.school_id is not None:
        user.school_id = user_data.school_id

    db.commit()
    db.refresh(user)
    return user


@router.delete("/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_system_admin)
):
    """删除用户（仅系统管理员）"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )

    # 不能删除自己
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="不能删除自己"
        )

    # 不能删除系统管理员
    if user.role == UserRole.SYSTEM_ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="不能删除系统管理员"
        )

    db.delete(user)
    db.commit()
    return None


class BatchStudentCreate(BaseModel):
    """批量创建学生"""
    students: List[dict]
    # 每个学生包含：username(必填), nickname(可选), password(可选), school_id(可选)
    # 如果不提供password，会自动生成
    auto_generate_password: bool = False
    # 如果不提供username，会自动生成
    auto_generate_username: bool = False


class BatchStudentResponse(BaseModel):
    """批量创建响应"""
    total: int
    success: int
    failed: int
    created_users: List[dict]


@router.post("/batch", response_model=BatchStudentResponse, status_code=status.HTTP_201_CREATED)
async def batch_create_students(
    data: BatchStudentCreate,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_school_admin_or_above)
):
    """批量创建学生"""
    # 学校管理员只能创建本校学生
    if current_user.role == UserRole.SCHOOL_ADMIN:
        # 检查所有学生的school_id是否在当前学校
        for student_data in data.students:
            if student_data.get('school_id') and student_data['school_id'] != current_user.school_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="只能为本校创建学生"
                )

    total = len(data.students)
    success = 0
    failed = 0
    created_users = []

    for student_data in data.students:
        try:
            # 自动生成学号
            if data.auto_generate_username or not student_data.get('username'):
                username = _generate_student_id()
            else:
                username = student_data['username']

            # 自动生成密码
            if data.auto_generate_password or not student_data.get('password'):
                password = _generate_password()
            else:
                password = student_data['password']

            # 检查用户名是否已存在
            existing_user = db.query(User).filter(User.username == username).first()
            if existing_user:
                failed += 1
                created_users.append({
                    'username': username,
                    'success': False,
                    'error': '用户名已存在'
                })
                continue

            # 创建用户
            from ..utils.security import get_password_hash
            new_user = User(
                username=username,
                password_hash=get_password_hash(password),
                nickname=student_data.get('nickname'),
                role=UserRole.STUDENT,
                school_id=student_data.get('school_id') or current_user.school_id
            )
            db.add(new_user)
            db.commit()
            db.refresh(new_user)

            success += 1
            created_users.append({
                'id': new_user.id,
                'username': new_user.username,
                'nickname': new_user.nickname,
                'password': password,  # 返回密码供前端显示
                'success': True
            })
        except Exception as e:
            failed += 1
            created_users.append({
                'username': student_data.get('username', 'unknown'),
                'success': False,
                'error': str(e)
            })
            db.rollback()

    return BatchStudentResponse(
        total=total,
        success=success,
        failed=failed,
        created_users=created_users
    )


class BatchUpdateSchoolRequest(BaseModel):
    """批量更新学校"""
    user_ids: List[int]
    school_id: int


class BatchUpdateResponse(BaseModel):
    """批量操作响应"""
    total: int
    success: int
    failed: int
    failed_users: List[dict]


@router.put("/batch/school", response_model=BatchUpdateResponse, status_code=status.HTTP_200_OK)
async def batch_update_school(
    request: BatchUpdateSchoolRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_system_admin)
):
    """批量设置学校（仅系统管理员）"""
    total = len(request.user_ids)
    success = 0
    failed = 0
    failed_users = []

    for user_id in request.user_ids:
        try:
            user = db.query(User).filter(User.id == user_id).first()
            if not user:
                failed += 1
                failed_users.append({
                    'user_id': user_id,
                    'error': '用户不存在'
                })
                continue

            # 不能修改系统管理员的学校
            if user.role == UserRole.SYSTEM_ADMIN:
                failed += 1
                failed_users.append({
                    'user_id': user_id,
                    'username': user.username,
                    'error': '不能修改系统管理员的学校'
                })
                continue

            user.school_id = request.school_id
            db.commit()
            db.refresh(user)

            success += 1
        except Exception as e:
            failed += 1
            failed_users.append({
                'user_id': user_id,
                'error': str(e)
            })
            db.rollback()

    return BatchUpdateResponse(
        total=total,
        success=success,
        failed=failed,
        failed_users=failed_users
    )


class BatchResetPasswordRequest(BaseModel):
    """批量重置密码"""
    user_ids: List[int]
    password: Optional[str] = None  # 如果不提供，自动生成密码


@router.put("/batch/reset-password", status_code=status.HTTP_200_OK)
async def batch_reset_password(
    request: BatchResetPasswordRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_system_admin)
):
    """批量重置密码（仅系统管理员），返回Excel文件"""
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    total = len(request.user_ids)
    results = []

    for user_id in request.user_ids:
        try:
            user = db.query(User).filter(User.id == user_id).first()
            if not user:
                results.append({
                    '用户ID': user_id,
                    '用户名': '-',
                    '姓名': '-',
                    '学校ID': '-',
                    '状态': '失败',
                    '错误信息': '用户不存在',
                    '新密码': '-'
                })
                continue

            # 不能重置系统管理员的密码
            if user.role == UserRole.SYSTEM_ADMIN:
                results.append({
                    '用户ID': user_id,
                    '用户名': user.username,
                    '姓名': user.nickname or '-',
                    '学校ID': user.school_id or '-',
                    '状态': '失败',
                    '错误信息': '不能重置系统管理员的密码',
                    '新密码': '-'
                })
                continue

            # 不能重置自己的密码
            if user_id == current_user.id:
                results.append({
                    '用户ID': user_id,
                    '用户名': user.username,
                    '姓名': user.nickname or '-',
                    '学校ID': user.school_id or '-',
                    '状态': '失败',
                    '错误信息': '不能重置自己的密码',
                    '新密码': '-'
                })
                continue

            # 生成或使用提供的密码
            password = request.password or _generate_password()
            from ..utils.security import get_password_hash
            user.password_hash = get_password_hash(password)
            db.commit()
            db.refresh(user)

            results.append({
                '用户ID': user_id,
                '用户名': user.username,
                '姓名': user.nickname or '-',
                '学校ID': user.school_id or '-',
                '状态': '成功',
                '错误信息': '-',
                '新密码': password
            })
        except Exception as e:
            results.append({
                '用户ID': user_id,
                '用户名': '-',
                '姓名': '-',
                '学校ID': '-',
                '状态': '失败',
                '错误信息': str(e),
                '新密码': '-'
            })
            db.rollback()

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "密码重置结果"

    # 添加表头
    headers = list(results[0].keys()) if results else ['用户ID', '用户名', '姓名', '学校ID', '状态', '错误信息', '新密码']
    ws.append(headers)

    # 添加数据
    for row in results:
        ws.append([row.get(header, '') for header in headers])

    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 统计信息
    success_count = len([r for r in results if r['状态'] == '成功'])
    failed_count = len([r for r in results if r['状态'] == '失败'])

    # 添加统计行
    ws.append([])
    ws.append(['总计', total])
    ws.append(['成功', success_count])
    ws.append(['失败', failed_count])
    ws.append(['操作人', f"{current_user.username} ({current_user.nickname or '-'})"])
    ws.append(['操作时间', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"password_reset_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


class BatchDeleteRequest(BaseModel):
    """批量删除用户"""
    user_ids: List[int]


@router.delete("/batch", response_model=BatchUpdateResponse, status_code=status.HTTP_200_OK)
async def batch_delete_users(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_system_admin)
):
    """批量删除用户（仅系统管理员）"""
    total = len(request.user_ids)
    success = 0
    failed = 0
    failed_users = []

    for user_id in request.user_ids:
        try:
            user = db.query(User).filter(User.id == user_id).first()
            if not user:
                failed += 1
                failed_users.append({
                    'user_id': user_id,
                    'error': '用户不存在'
                })
                continue

            # 不能删除自己
            if user_id == current_user.id:
                failed += 1
                failed_users.append({
                    'user_id': user_id,
                    'username': user.username,
                    'error': '不能删除自己'
                })
                continue

            # 不能删除系统管理员
            if user.role == UserRole.SYSTEM_ADMIN:
                failed += 1
                failed_users.append({
                    'user_id': user_id,
                    'username': user.username,
                    'error': '不能删除系统管理员'
                })
                continue

            db.delete(user)
            db.commit()

            success += 1
        except Exception as e:
            failed += 1
            failed_users.append({
                'user_id': user_id,
                'error': str(e)
            })
            db.rollback()

    return BatchUpdateResponse(
        total=total,
        success=success,
        failed=failed,
        failed_users=failed_users
    )


@router.post("/import", status_code=status.HTTP_201_CREATED)
async def import_students(
    db: Session = Depends(get_db),
    current_user: CurrentUserResponse = Depends(require_school_admin_or_above)
):
    """导入学生名单（从上传的Excel文件）"""
    from fastapi import UploadFile, File

    # 由于需要读取上传的文件，这里应该从请求中获取
    # 前端需要通过FormData上传Excel文件
    # 这个端点需要前端传递文件，这里只是定义接口
    raise HTTPException(
        status_code=status.HTTP_501_NOT_IMPLEMENTED,
        detail="请使用前端导入功能上传Excel文件"
    )


def _generate_student_id() -> str:
    """生成学号（格式：2024XXXX）"""
    year = 2024
    random_num = random.randint(1000, 9999)
    return f"{year}{random_num}"


def _generate_password() -> str:
    """生成随机密码（8位，包含字母和数字）"""
    chars = string.ascii_letters + string.digits
    password = ''.join(random.choice(chars) for _ in range(8))
    return password
